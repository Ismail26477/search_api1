import openpyxl
import os
from django.http import HttpResponse
from django.shortcuts import render
from .models import FAQ
from .forms import SearchForm
from openpyxl.drawing.image import Image as XLImage


def question(request):
    faqs = []

    if request.method == 'GET':
        form = SearchForm(request.GET)
        
        if form.is_valid():
            search_query = form.cleaned_data['search_query']
            print(f"Search Query: {search_query}")  # Debugging line
            
            faqs = FAQ.objects.filter(question__icontains=search_query)
            print(f"Matching FAQs: {faqs}")  # Debugging line

            # Export to Excel if faqs found
            if faqs:
                export_to_excel(faqs)

    else:
        form = SearchForm()

    context = {
        'form': form,
        'faqs': faqs,
    }
    return render(request, 'question.html', context)

def export_to_excel(faqs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FAQs"

    headers = ["Question", "Answer", "Reference", "Image"]
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    for row_num, faq in enumerate(faqs, 2):
        ws[f"A{row_num}"] = faq.question
        ws[f"B{row_num}"] = faq.answer
        ws[f"C{row_num}"] = faq.reference

        if faq.image:
            img = XLImage(faq.image.path)
            img.width = 50
            img.height = 50
            ws.add_image(img, f"D{row_num}")

    wb.save("C:\\DjangoProject\\MY_Project\\abc.xlsx")
